from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from iss_parser_core.iss_parser import IssParseResult


@dataclass
class SampleFeederRow:
    component_name: str
    sample_feeder_value: str
    sample_package_value: str
    source_path: str
    source_row: int


class FeederMappingAnalyzer:
    def load_sample_rows(self, sample_path: str | Path) -> list[SampleFeederRow]:
        path = Path(sample_path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._load_from_csv(path)
        if suffix == ".xlsx":
            return self._load_from_xlsx(path)
        raise ValueError(f"不支持的样品文件类型: {path.name}")

    def build_detail_rows(
        self,
        parse_result: IssParseResult,
        sample_rows: list[SampleFeederRow],
        component_map: dict[str, Any],
        feeders_by_component: dict[str, list[Any]],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for index, sample in enumerate(sample_rows, start=1):
            component = component_map.get(sample.component_name)
            component_extra = getattr(component, "extra", {}) or {}
            feeders = feeders_by_component.get(sample.component_name, [])

            if feeders:
                for feeder in feeders:
                    rows.append(
                        {
                            "序号": index,
                            "componentName": sample.component_name,
                            "样品飞达值": sample.sample_feeder_value,
                            "样品包装值": sample.sample_package_value,
                            "package": component_extra.get("package", ""),
                            "reelTypeId": component_extra.get("feeder.reelTypeId", ""),
                            "pitch": component_extra.get("feederPitch.pitch", ""),
                            "pitchCount": component_extra.get("feederPitch.count", ""),
                            "feeder.typeId": feeder.feeder_type,
                            "bankKind": feeder.bank_kind,
                            "stationId": feeder.station_id,
                            "bankPos": feeder.bank_pos,
                            "holeNo": feeder.hole_no,
                            "lane": feeder.lane,
                            "componentType": component_extra.get("componentType", ""),
                            "packageCode": getattr(component, "package_code", "") if component else "",
                            "样品来源": sample.source_path,
                            "样品行号": sample.source_row,
                        }
                    )
            else:
                rows.append(
                    {
                        "序号": index,
                        "componentName": sample.component_name,
                        "样品飞达值": sample.sample_feeder_value,
                        "样品包装值": sample.sample_package_value,
                        "package": component_extra.get("package", ""),
                        "reelTypeId": component_extra.get("feeder.reelTypeId", ""),
                        "pitch": component_extra.get("feederPitch.pitch", ""),
                        "pitchCount": component_extra.get("feederPitch.count", ""),
                        "feeder.typeId": "",
                        "bankKind": "",
                        "stationId": "",
                        "bankPos": "",
                        "holeNo": "",
                        "lane": "",
                        "componentType": component_extra.get("componentType", ""),
                        "packageCode": getattr(component, "package_code", "") if component else "",
                        "样品来源": sample.source_path,
                        "样品行号": sample.source_row,
                    }
                )
        return rows

    def build_candidate_summary(self, detail_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        grouped: dict[str, dict[str, Any]] = {}
        for row in detail_rows:
            feeder_value = str(row.get("样品飞达值", "")).strip()
            combo = (
                str(row.get("package", "")).strip(),
                str(row.get("reelTypeId", "")).strip(),
                str(row.get("pitch", "")).strip(),
                str(row.get("pitchCount", "")).strip(),
                str(row.get("feeder.typeId", "")).strip(),
                str(row.get("bankKind", "")).strip(),
                str(row.get("componentType", "")).strip(),
            )
            bucket = grouped.setdefault(
                feeder_value,
                {
                    "样品飞达值": feeder_value,
                    "组合集合": {},
                },
            )
            combo_key = "|".join(combo)
            combo_bucket = bucket["组合集合"].setdefault(
                combo_key,
                {
                    "package": combo[0],
                    "reelTypeId": combo[1],
                    "pitch": combo[2],
                    "pitchCount": combo[3],
                    "feeder.typeId": combo[4],
                    "bankKind": combo[5],
                    "componentType": combo[6],
                    "stationIds": set(),
                    "bankPosSet": set(),
                    "holeNoSet": set(),
                    "laneSet": set(),
                    "componentNames": set(),
                },
            )
            combo_bucket["componentNames"].add(str(row.get("componentName", "")).strip())
            combo_bucket["stationIds"].add(str(row.get("stationId", "")).strip())
            combo_bucket["bankPosSet"].add(str(row.get("bankPos", "")).strip())
            combo_bucket["holeNoSet"].add(str(row.get("holeNo", "")).strip())
            combo_bucket["laneSet"].add(str(row.get("lane", "")).strip())

        summary_rows: list[dict[str, Any]] = []
        summary_index = 1
        for feeder_value, bucket in grouped.items():
            combo_count = len(bucket["组合集合"])
            for combo_key, combo_bucket in bucket["组合集合"].items():
                summary_rows.append(
                    {
                        "序号": summary_index,
                        "样品飞达值": feeder_value,
                        "候选组合数量": combo_count,
                        "是否唯一": "是" if combo_count == 1 else "否",
                        "package": combo_bucket["package"],
                        "reelTypeId": combo_bucket["reelTypeId"],
                        "pitch": combo_bucket["pitch"],
                        "pitchCount": combo_bucket["pitchCount"],
                        "feeder.typeId": combo_bucket["feeder.typeId"],
                        "bankKind": combo_bucket["bankKind"],
                        "componentType": combo_bucket["componentType"],
                        "stationId集合": ",".join(sorted(item for item in combo_bucket["stationIds"] if item)),
                        "bankPos集合": ",".join(sorted(item for item in combo_bucket["bankPosSet"] if item)),
                        "holeNo集合": ",".join(sorted(item for item in combo_bucket["holeNoSet"] if item)),
                        "lane集合": ",".join(sorted(item for item in combo_bucket["laneSet"] if item)),
                        "componentName示例": ",".join(sorted(combo_bucket["componentNames"])),
                    }
                )
                summary_index += 1
        return summary_rows

    def export_report(
        self,
        output_path: str | Path,
        detail_rows: list[dict[str, Any]],
        summary_rows: list[dict[str, Any]],
    ) -> Path:
        path = Path(output_path).resolve()
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)

        for sheet_name, rows in [
            ("飞达映射明细", detail_rows),
            ("飞达候选映射验证", summary_rows),
        ]:
            ws = workbook.create_sheet(sheet_name)
            self._write_rows(ws, rows)

        workbook.save(path)
        return path

    def _write_rows(self, worksheet, rows: list[dict[str, Any]]) -> None:
        if not rows:
            worksheet.append(["无数据"])
            return
        headers = list(rows[0].keys())
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    def _load_from_csv(self, path: Path) -> list[SampleFeederRow]:
        rows = list(csv.reader(path.open("r", encoding="gbk", errors="replace")))
        header_index = self._find_csv_header(rows)
        if header_index < 0:
            raise ValueError(f"未找到样品表头: {path.name}")
        header = rows[header_index]
        index_map = {name: idx for idx, name in enumerate(header)}
        component_index = self._find_index(index_map, ["元件名", "物料编码"])
        feeder_index = self._find_index(index_map, ["类型", "飞达", "送料器类型", "供料器类型"])
        package_index = self._find_index(index_map, ["包装"])
        result: list[SampleFeederRow] = []
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            component_name = self._safe_get(row, component_index)
            feeder_value = self._safe_get(row, feeder_index)
            package_value = self._safe_get(row, package_index)
            if not component_name or not feeder_value:
                continue
            if component_name in {"元件名", "物料编码"} or feeder_value in {"类型", "飞达", "送料器类型", "供料器类型"}:
                continue
            result.append(
                SampleFeederRow(
                    component_name=component_name,
                    sample_feeder_value=feeder_value,
                    sample_package_value=package_value,
                    source_path=str(path),
                    source_row=row_number,
                )
            )
        return result

    def _load_from_xlsx(self, path: Path) -> list[SampleFeederRow]:
        wb = load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        values = list(ws.iter_rows(values_only=True))
        header_index = self._find_xlsx_header(values)
        if header_index < 0:
            raise ValueError(f"未找到样品表头: {path.name}")
        header_row = [str(item or "").strip() for item in values[header_index]]
        index_map = {name: idx for idx, name in enumerate(header_row) if name}
        component_index = self._find_index(index_map, ["元件名", "物料编码"])
        feeder_index = self._find_index(index_map, ["类型", "飞达", "送料器类型", "供料器类型"])
        package_index = self._find_index(index_map, ["包装", "封装"])
        result: list[SampleFeederRow] = []
        for row_number, row in enumerate(values[header_index + 1 :], start=header_index + 2):
            component_name = self._safe_get_seq(row, component_index)
            feeder_value = self._safe_get_seq(row, feeder_index)
            package_value = self._safe_get_seq(row, package_index)
            if not component_name or not feeder_value:
                continue
            if component_name in {"元件名", "物料编码"} or feeder_value in {"类型", "飞达", "送料器类型", "供料器类型"}:
                continue
            result.append(
                SampleFeederRow(
                    component_name=component_name,
                    sample_feeder_value=feeder_value,
                    sample_package_value=package_value,
                    source_path=str(path),
                    source_row=row_number,
                )
            )
        return result

    def _find_csv_header(self, rows: list[list[str]]) -> int:
        for index, row in enumerate(rows):
            joined = ",".join(row)
            if "元件名" in joined and ("类型" in joined or "飞达" in joined or "送料器类型" in joined):
                return index
        return -1

    def _find_xlsx_header(self, rows: list[tuple[Any, ...]]) -> int:
        for index, row in enumerate(rows):
            values = [str(item or "").strip() for item in row]
            if ("元件名" in values or "物料编码" in values) and ("类型" in values or "飞达" in values or "送料器类型" in values):
                return index
        return -1

    def _find_index(self, index_map: dict[str, int], candidates: list[str]) -> int:
        for name in candidates:
            if name in index_map:
                return index_map[name]
        return -1

    def _safe_get(self, row: list[str], index: int) -> str:
        if index < 0 or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    def _safe_get_seq(self, row: tuple[Any, ...], index: int) -> str:
        if index < 0 or index >= len(row):
            return ""
        return str(row[index] or "").strip()
