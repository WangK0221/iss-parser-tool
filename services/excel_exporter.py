from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import CUSTOMER_EXPORT_WIDTHS, DATETIME_FORMAT, MAIN_SHEET_NAMES, RAW_SHEET_NAMES


class ExcelExporter:
    """负责将结构化结果导出为 Excel 文件。"""

    def export(self, output_dir: Path, file_stem: str, sheets: dict[str, list[dict[str, Any]]]) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)

        for sheet_key, rows in sheets.items():
            sheet_title = self._resolve_sheet_title(sheet_key)
            worksheet = workbook.create_sheet(title=sheet_title[:31])
            self._write_sheet(worksheet, rows)

        timestamp = datetime.now().strftime(DATETIME_FORMAT)
        output_path = output_dir / f"{file_stem}_{timestamp}.xlsx"
        workbook.save(output_path)
        return output_path

    def _resolve_sheet_title(self, sheet_key: str) -> str:
        if sheet_key in MAIN_SHEET_NAMES:
            return MAIN_SHEET_NAMES[sheet_key]
        if sheet_key.endswith("_raw"):
            if sheet_key.startswith("components"):
                return RAW_SHEET_NAMES["components"]
            if sheet_key.startswith("placements"):
                return RAW_SHEET_NAMES["placements"]
            if sheet_key.startswith("feeders"):
                return RAW_SHEET_NAMES["feeders"]
        return sheet_key

    def _write_sheet(self, worksheet, rows: Any) -> None:
        if isinstance(rows, dict) and rows.get("sheet_type") == "customer_format":
            self._write_customer_sheet(worksheet, rows)
            return

        if not rows:
            worksheet.append(["无数据"])
            worksheet.freeze_panes = "A2"
            return

        headers = list(rows[0].keys())
        worksheet.append(headers)
        for row in rows:
            worksheet.append([self._normalize_cell_value(row.get(header, "")) for header in headers])

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        worksheet.freeze_panes = "A2"
        self._auto_fit(worksheet)

    def _write_customer_sheet(self, worksheet, sheet_data: dict[str, Any]) -> None:
        title = self._stringify(sheet_data.get("title", ""))
        fields = list(sheet_data.get("fields", []))
        header_map = dict(sheet_data.get("headers", {}))
        rows = list(sheet_data.get("rows", []))
        sections = list(sheet_data.get("sections", []))
        layout = dict(sheet_data.get("layout", {}))
        detected_line_name = self._stringify(sheet_data.get("detected_line_name", ""))

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        plain_fill = PatternFill(fill_type=None)
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")

        if sections:
            current_row = 1
            block_index = 0
            split_by_station_group = bool(layout.get("split_by_station_group", True))
            for section in sections:
                groups = list(section.get("groups", [])) if split_by_station_group else [{"group_key": "ALL", "rows": list(section.get("rows", []))}]
                for group in groups:
                    if not group.get("rows"):
                        continue
                    if block_index > 0:
                        current_row += 1
                        worksheet.row_breaks.append(Break(id=current_row - 1))
                    current_row = self._write_customer_block(
                        worksheet=worksheet,
                        start_row=current_row,
                        title=title,
                        fields=fields,
                        header_map=header_map,
                        rows=group["rows"],
                        machine_no=str(section.get("machine_no", "")).strip(),
                        layout=layout,
                        detected_line_name=detected_line_name,
                        border=border,
                        fill=plain_fill,
                        header_font=header_font,
                        center=center,
                    )
                    block_index += 1
            worksheet.freeze_panes = "A4"
        else:
            self._write_customer_block(
                worksheet=worksheet,
                start_row=1,
                title=title,
                fields=fields,
                header_map=header_map,
                rows=rows,
                machine_no="",
                layout=layout,
                detected_line_name=detected_line_name,
                border=border,
                fill=plain_fill,
                header_font=header_font,
                center=center,
            )
            worksheet.freeze_panes = "A4"

        self._apply_customer_widths(worksheet, fields)

    def _write_customer_block(
        self,
        worksheet,
        start_row: int,
        title: str,
        fields: list[str],
        header_map: dict[str, str],
        rows: list[dict[str, Any]],
        machine_no: str,
        layout: dict[str, Any],
        detected_line_name: str,
        border,
        fill,
        header_font,
        center,
    ) -> int:
        if fields:
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(fields))
        title_cell = worksheet.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center
        self._apply_row_style(worksheet, start_row, len(fields), border, fill)
        title_cell.border = border
        title_cell.fill = fill

        line_label = self._stringify(layout.get("line_label", "线体"))
        line_value = self._stringify(layout.get("line_value", "")).strip() or detected_line_name
        machine_label = self._stringify(layout.get("machine_label", "机台"))
        machine_value = f"{machine_no}号机" if machine_no else "未知"

        meta_row = start_row + 1
        self._apply_row_style(worksheet, meta_row, len(fields), border, fill)
        meta_entries = [
            f"{line_label}：{line_value}",
            f"{machine_label}：{machine_value}",
        ]
        for index, value in enumerate(meta_entries, start=1):
            cell = worksheet.cell(row=meta_row, column=index, value=value)
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
            cell.fill = fill

        header_row = start_row + 2
        self._write_customer_headers(worksheet, header_row, fields, header_map, header_font, center, border, fill)

        current_row = header_row + 1
        for row in rows:
            self._write_customer_row(worksheet, current_row, fields, row, border, center)
            current_row += 1

        return current_row

    def _write_customer_headers(self, worksheet, row_index: int, fields: list[str], header_map: dict[str, str], header_font, center, border, fill) -> None:
        for column_index, field in enumerate(fields, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=header_map.get(field, field))
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            cell.fill = fill

    def _write_customer_row(self, worksheet, row_index: int, fields: list[str], row: dict[str, Any], border, center) -> None:
        for column_index, field in enumerate(fields, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=self._normalize_cell_value(row.get(field, "")))
            cell.border = border
            if field in {"machine_station", "station", "package", "feeder", "quantity"}:
                cell.alignment = center

    def _apply_customer_widths(self, worksheet, fields: list[str]) -> None:
        for index, field in enumerate(fields, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = CUSTOMER_EXPORT_WIDTHS.get(field, 16)

    def _apply_row_style(self, worksheet, row_index: int, column_count: int, border, fill) -> None:
        for column_index in range(1, column_count + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = border
            cell.fill = fill

    def _auto_fit(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                max_length = max(max_length, len(self._stringify(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 60)

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, dict):
            return "; ".join(f"{key}={val}" for key, val in value.items())
        return str(value)

    def _normalize_cell_value(self, value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, dict):
            return "; ".join(f"{key}={val}" for key, val in value.items())
        return value
